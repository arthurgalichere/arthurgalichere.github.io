import json
import os
import sys
import tempfile
import urllib.error
import urllib.request
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl


DEFAULT_EXCEL_URL = (
    "https://www.dropbox.com/scl/fi/z0dbe74ywv0ws3yw4l8gt/"
    "CV_Arthur_Galichere_excel.xlsx"
    "?rlkey=l736567qyln1ql0s7ws2nz21q&st=vx9y8fjx&dl=1"
)

EXPECTED_SECTIONS = {
    "Employment",
    "Professional Development",
    "Conference Paper Reviewer",
    "Teaching Awards and Qualifications",
    "Teaching Experience",
    "Additional Teaching and Supervisory Experience",
    "Academic Leadership and Development",
    "Administrative and Collegial Experience",
    "Research Presentations",
}

CV_FIELDS = (
    "role",
    "institution",
    "date",
    "details",
    "category",
    "other_details",
)

PAPER_FIELDS = (
    "title",
    "category",
    "journal",
    "date",
    "coauthors",
    "abstract",
    "url",
    "status",
)

PAPER_CATEGORIES = {
    "working paper": "Working Papers",
    "working papers": "Working Papers",
    "work in progress": "Work in Progress",
    "published": "Published",
    "publication": "Published",
    "publications": "Published",
}


def cell_text(value):
    return "" if value is None else str(value).strip()


def normalized_header(value):
    return cell_text(value).casefold().replace(" ", "_")


def header_map(sheet):
    return {
        normalized_header(sheet.cell(row=2, column=column).value): column
        for column in range(1, sheet.max_column + 1)
        if cell_text(sheet.cell(row=2, column=column).value)
    }


def row_values(sheet, row, fields, columns):
    return {
        field: cell_text(sheet.cell(row=row, column=columns[field]).value)
        if columns.get(field)
        else ""
        for field in fields
    }


def download_workbook(url):
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "Arthur-Galichere-CV-Updater/2.0"},
    )

    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            content = response.read()
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        raise RuntimeError(f"Could not download the Excel workbook: {exc}") from exc

    if len(content) < 1_000:
        raise RuntimeError("The downloaded Excel workbook is unexpectedly small.")

    try:
        return openpyxl.load_workbook(
            BytesIO(content),
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise RuntimeError(
            f"The downloaded file is not a readable Excel workbook: {exc}"
        ) from exc


def is_papers_sheet(sheet):
    sheet_name = cell_text(sheet.title).casefold()
    section_title = cell_text(sheet.cell(row=1, column=1).value).casefold()
    return sheet_name == "p" or section_title in {"papers", "research papers"}


def read_cv_sheet(sheet):
    section_title = cell_text(sheet.cell(row=1, column=1).value) or sheet.title
    headers = header_map(sheet)
    columns = {field: headers.get(field) for field in CV_FIELDS}

    if not columns["role"]:
        raise ValueError(f"Worksheet '{sheet.title}' has no 'role' column.")

    items = []
    for row in range(3, sheet.max_row + 1):
        values = row_values(sheet, row, CV_FIELDS, columns)

        if not any(values.values()):
            continue

        details = values["details"]
        if values["other_details"]:
            details = "\n".join(filter(None, (details, values["other_details"])))

        items.append(
            {
                "role": values["role"],
                "institution": values["institution"],
                "date": values["date"],
                "details": details,
                "category": values["category"],
            }
        )

    if not items:
        raise ValueError(f"Worksheet '{sheet.title}' contains no CV records.")

    return section_title, items


def canonical_paper_category(value):
    category = cell_text(value)
    canonical = PAPER_CATEGORIES.get(category.casefold())
    if not canonical:
        allowed = ", ".join(sorted(set(PAPER_CATEGORIES.values())))
        raise ValueError(
            f"Unknown paper category '{category}'. Expected one of: {allowed}."
        )
    return canonical


def read_papers_sheet(sheet):
    headers = header_map(sheet)
    columns = {field: headers.get(field) for field in PAPER_FIELDS}

    missing_headers = [
        field for field in ("title", "category") if not columns.get(field)
    ]
    if missing_headers:
        raise ValueError(
            f"Papers worksheet '{sheet.title}' is missing required columns: "
            + ", ".join(missing_headers)
        )

    papers = []
    for row in range(3, sheet.max_row + 1):
        values = row_values(sheet, row, PAPER_FIELDS, columns)

        if not any(values.values()):
            continue

        if not values["title"]:
            raise ValueError(
                f"Papers worksheet '{sheet.title}', row {row}, has no title."
            )
        if not values["category"]:
            raise ValueError(
                f"Papers worksheet '{sheet.title}', row {row}, has no category."
            )

        papers.append(
            {
                "title": values["title"],
                "category": canonical_paper_category(values["category"]),
                "journal": values["journal"],
                "date": values["date"],
                "coauthors": values["coauthors"],
                "abstract": values["abstract"],
                "url": values["url"],
                "status": values["status"],
            }
        )

    if not papers:
        raise ValueError(f"Papers worksheet '{sheet.title}' contains no papers.")

    return papers


def group_teaching_experience(section_title, items):
    institutions = {}

    for item in items:
        institution = item.get("institution") or "Other"
        category = item.get("category") or "General"
        institutions.setdefault(institution, {}).setdefault(category, []).append(item)

    subsections = []
    for institution, categories in institutions.items():
        institution_items = []
        for category, category_items in categories.items():
            institution_items.append({"isFormatHeader": True, "role": category})
            institution_items.extend(category_items)
        subsections.append({"title": institution, "items": institution_items})

    return {"title": section_title, "subsections": subsections}


def group_additional_teaching(section_title, items):
    categories = {}

    for item in items:
        category = item.get("category") or "General"
        categories.setdefault(category, []).append(item)

    return {
        "title": section_title,
        "subsections": [
            {"title": category, "items": category_items}
            for category, category_items in categories.items()
        ],
    }


def build_data(workbook):
    sections = []
    papers = None

    for sheet in workbook.worksheets:
        if is_papers_sheet(sheet):
            if papers is not None:
                raise ValueError("The workbook contains more than one papers worksheet.")
            papers = read_papers_sheet(sheet)
            continue

        section_title, items = read_cv_sheet(sheet)
        section_lower = section_title.casefold()

        if section_lower == "teaching experience":
            section = group_teaching_experience(section_title, items)
        elif section_lower == "additional teaching and supervisory experience":
            section = group_additional_teaching(section_title, items)
        else:
            section = {"title": section_title, "items": items}

        sections.append(section)

    found_sections = {section["title"] for section in sections}
    missing_sections = EXPECTED_SECTIONS - found_sections
    if missing_sections:
        missing = ", ".join(sorted(missing_sections))
        raise ValueError(f"The workbook is missing expected CV sections: {missing}")

    unexpected_sections = found_sections - EXPECTED_SECTIONS
    if unexpected_sections:
        unexpected = ", ".join(sorted(unexpected_sections))
        raise ValueError(f"The workbook contains unexpected CV sections: {unexpected}")

    if papers is None:
        raise ValueError(
            "The workbook has no papers worksheet. Add a worksheet named 'P' "
            "with 'Papers' or 'Research Papers' in cell A1."
        )

    return sections, papers


def read_previous_data(output_path):
    if not output_path.is_file():
        return None

    try:
        with output_path.open(encoding="utf-8") as file:
            data = json.load(file)
    except (OSError, json.JSONDecodeError):
        return None

    return data if isinstance(data, dict) else None


def choose_update_date(new_sections, new_papers, previous_data):
    if previous_data:
        content_unchanged = (
            previous_data.get("sections") == new_sections
            and previous_data.get("papers") == new_papers
        )
        previous_date = previous_data.get("last_updated")

        if content_unchanged:
            try:
                date.fromisoformat(previous_date)
                return previous_date
            except (TypeError, ValueError):
                pass

    return date.today().isoformat()


def validate_items(items, location):
    if not isinstance(items, list) or not items:
        raise ValueError(f"{location} contains no items.")

    for index, item in enumerate(items, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"{location}, item {index}, must be an object.")

        if item.get("isFormatHeader"):
            if not cell_text(item.get("role")):
                raise ValueError(
                    f"{location}, format header {index}, has no role text."
                )
            continue

        if not any(
            cell_text(item.get(field))
            for field in ("role", "institution", "date", "details")
        ):
            raise ValueError(f"{location}, item {index}, is empty.")


def validate_cv_data(data):
    if not isinstance(data, dict):
        raise ValueError("CV data must be a JSON object.")

    try:
        date.fromisoformat(data.get("last_updated", ""))
    except (TypeError, ValueError) as exc:
        raise ValueError(
            "CV data must contain a valid ISO 'last_updated' date."
        ) from exc

    sections = data.get("sections")
    if not isinstance(sections, list) or not sections:
        raise ValueError("CV data must contain a non-empty 'sections' list.")

    found_sections = set()
    for section in sections:
        if not isinstance(section, dict) or not cell_text(section.get("title")):
            raise ValueError("Every CV section must be an object with a title.")

        title = cell_text(section["title"])
        found_sections.add(title)

        if "subsections" in section:
            subsections = section["subsections"]
            if not isinstance(subsections, list) or not subsections:
                raise ValueError(f"Section '{title}' contains no subsections.")

            for subsection in subsections:
                if not isinstance(subsection, dict) or not cell_text(
                    subsection.get("title")
                ):
                    raise ValueError(
                        f"Every subsection in '{title}' must have a title."
                    )
                validate_items(
                    subsection.get("items"),
                    f"Section '{title}', subsection '{subsection['title']}'",
                )
        else:
            validate_items(section.get("items"), f"Section '{title}'")

    missing_sections = EXPECTED_SECTIONS - found_sections
    if missing_sections:
        missing = ", ".join(sorted(missing_sections))
        raise ValueError(f"CV data is missing expected sections: {missing}")

    papers = data.get("papers")
    if not isinstance(papers, list) or not papers:
        raise ValueError("CV data must contain a non-empty 'papers' list.")

    allowed_categories = set(PAPER_CATEGORIES.values())
    for index, paper in enumerate(papers, start=1):
        if not isinstance(paper, dict):
            raise ValueError(f"Paper {index} must be an object.")

        title = cell_text(paper.get("title"))
        category = cell_text(paper.get("category"))
        if not title:
            raise ValueError(f"Paper {index} has no title.")
        if category not in allowed_categories:
            allowed = ", ".join(sorted(allowed_categories))
            raise ValueError(
                f"Paper '{title}' has invalid category '{category}'. "
                f"Expected one of: {allowed}."
            )


def write_json_atomically(data, output_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)

    temporary_path = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=output_path.parent,
            prefix=f".{output_path.name}.",
            suffix=".tmp",
            delete=False,
        ) as temporary_file:
            json.dump(data, temporary_file, indent=2, ensure_ascii=False)
            temporary_file.write("\n")
            temporary_path = Path(temporary_file.name)

        with temporary_path.open(encoding="utf-8") as file:
            validate_cv_data(json.load(file))

        os.replace(temporary_path, output_path)
    finally:
        if temporary_path and temporary_path.exists():
            temporary_path.unlink()


def main():
    base_dir = Path(__file__).resolve().parent
    output_path = base_dir / "cv.json"
    excel_url = os.environ.get("CV_EXCEL_URL", DEFAULT_EXCEL_URL).strip()

    if not excel_url:
        raise RuntimeError("CV_EXCEL_URL is empty.")

    previous_data = read_previous_data(output_path)
    workbook = download_workbook(excel_url)

    try:
        sections, papers = build_data(workbook)
    finally:
        workbook.close()

    cv_data = {
        "last_updated": choose_update_date(sections, papers, previous_data),
        "sections": sections,
        "papers": papers,
    }

    validate_cv_data(cv_data)
    write_json_atomically(cv_data, output_path)

    print(
        f"Updated {output_path} with {len(sections)} CV sections and "
        f"{len(papers)} papers; information date: {cv_data['last_updated']}."
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"CV extraction failed: {exc}", file=sys.stderr)
        sys.exit(1)
