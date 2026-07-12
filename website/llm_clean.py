import os
import sys
import json
import urllib.request
import urllib.error
import openpyxl
import tempfile
import time

def clean_json_with_llm():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    temp_text_path = os.path.join(base_dir, "raw_cv.txt")
    output_json_path = os.path.join(base_dir, "cv.json")
    
    merged_sections = []
    
    # --- 1. PDF Extraction (With Retry Logic) ---
    if os.path.exists(temp_text_path):
        with open(temp_text_path, "r", encoding="utf-8") as f:
            raw_text = f.read()
        
        api_key = os.environ.get("GEMINI_API_KEY")
        if api_key:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
            payload = {
                "contents": [{"parts": [{"text": f"Extract 'Employment' and 'Research Presentations' from this CV. Output ONLY as JSON.\n{raw_text}"}]}],
                "generationConfig": {"responseMimeType": "application/json", "temperature": 0.1}
            }
            
            # Retry loop for PDF extraction
            for attempt in range(3):
                try:
                    req = urllib.request.Request(url, data=json.dumps(payload).encode("utf-8"), headers={"Content-Type": "application/json"}, method="POST")
                    with urllib.request.urlopen(req) as response:
                        result = json.loads(response.read().decode("utf-8"))
                        json_text_out = result["candidates"][0]["content"]["parts"][0]["text"]
                        gemini_data = json.loads(json_text_out)
                        if "sections" in gemini_data:
                            merged_sections.extend(gemini_data["sections"])
                            print(f"DEBUG: PDF extraction success. Added {len(gemini_data['sections'])} sections.")
                            break 
                except Exception as e:
                    print(f"DEBUG ERROR: PDF extraction failed attempt {attempt+1}: {e}")
                    time.sleep(16)
        else:
            print("DEBUG: No API Key found.")

    # --- 2. Excel Parsing ---
    excel_url = "https://www.dropbox.com/scl/fi/z0dbe74ywv0ws3yw4l8gt/CV_Arthur_Galichere_excel.xlsx?rlkey=l736567qyln1ql0s7ws2nz21q&st=vx9y8fjx&dl=1"
    
    # Define which sections require University/Category grouping
    GROUPED_SECTIONS = ["teaching experience", "additional teaching and supervisory experience"]

    try:
        with urllib.request.urlopen(excel_url) as response:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(response.read())
                temp_excel_path = tmp.name
        
        wb = openpyxl.load_workbook(temp_excel_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            section_title = str(sheet.cell(row=1, column=1).value or sheet_name).strip()
            
            # Identify columns
            header_map = {str(sheet.cell(row=2, column=col).value).strip().lower(): col 
                         for col in range(1, sheet.max_column + 1) if sheet.cell(row=2, column=col).value}
            
            cols = ["role", "institution", "date", "category", "details", "other_details"]
            col_map = {k: header_map.get(k) for k in cols}
            
            items = []
            for r in range(3, sheet.max_row + 1):
                # Safely get row values
                row_vals = {k: str(sheet.cell(row=r, column=col_map[k]).value or "").strip() 
                           if col_map.get(k) else "" for k in cols}
                
                if any(row_vals.values()):
                    # Combine details and other_details safely
                    combined_details = row_vals["details"]
                    if row_vals.get("other_details"):
                        combined_details = f"{combined_details}\n{row_vals['other_details']}".strip()
                    
                    items.append({
                        "role": row_vals["role"], 
                        "institution": row_vals["institution"], 
                        "date": row_vals["date"], 
                        "details": combined_details,
                        "category": row_vals.get("category", "")
                    })
            
            if items:
                # Check if this section needs grouping logic
                if section_title.lower() in GROUPED_SECTIONS:
                    by_inst = {}
                    for item in items:
                        inst = item.get("institution") or "Other"
                        category = item.get("category") or "General"
                        
                        if inst not in by_inst: by_inst[inst] = {}
                        if category not in by_inst[inst]: by_inst[inst][category] = []
                        
                        by_inst[inst][category].append(item)
                    
                    subsections = []
                    for inst, categories in by_inst.items():
                        inst_items = []
                        for cat, cat_items in categories.items():
                            inst_items.append({"isFormatHeader": True, "role": cat})
                            inst_items.extend(cat_items)
                        subsections.append({"title": inst, "items": inst_items})
                    
                    merged_sections.append({"title": section_title, "subsections": subsections})
                else:
                    # Append as standard flat section
                    merged_sections.append({"title": section_title, "items": items})
        
        if os.path.exists(temp_excel_path): os.remove(temp_excel_path)
    except Exception as e:
        print(f"DEBUG ERROR: Excel parsing failed: {e}")

    # --- 3. Save ---
    print(f"DEBUG: Saving {len(merged_sections)} total sections.")
    with open(output_json_path, "w", encoding="utf-8") as f:
        json.dump({"sections": merged_sections}, f, indent=2, ensure_ascii=False)
    print("Pipeline Complete.")

if __name__ == "__main__":
    clean_json_with_llm()
